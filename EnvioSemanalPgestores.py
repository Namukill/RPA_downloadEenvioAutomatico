import pandas as pd # Para manipulação de dados
from datetime import datetime # Para manipulação de datas
from openpyxl.utils import get_column_letter # Para manipulação de colunas do Excel
from openpyxl import load_workbook # Para manipulação de arquivos Excel
import os # Para manipulação de arquivos e diretórios
import pyautogui # Para automação de interface gráfica
import time # Para criar delays no código
import win32com.client as win32 # Para enviar e-mails via Outlook
import pymsgbox # Para exibir mensagens de alerta

def processar_planilha(): # Primeira função que processa a planilha Excel para ajustar tamanhos de colunas e filtrar por mês atual
    try:
        nome_arquivo = None
        df = pd.read_excel(
            r'' # Coloque o caminho do arquivo Excel base aqui 
        )
        print("Planilha carregada.")

        # Lista com o tamanho de cada coluna para ajustar, pois o dataframe salvo no Excel não possui o tamanho correto das colunas
        colunas_para_ajustar = {
            # coloque o nome da coluna entre aspas e o tamanho desejado em seguida
            '': 5, 
            '': 20,
            '': 25,
        }

        df['Data'] = pd.to_datetime(df['Data'], errors='coerce')# Converter coluna Data para formato dd/mm/yyyy
        hoje = datetime.today()

        df['Data'] = df['Data'].dt.date 
        df['Data'] = df['Data'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notnull(x) else '') # Pega a data no formato dd/mm/yyyy para filtro e para salvar no excel

        df['Data_dt'] = pd.to_datetime(df['Data'], format='%d/%m/%Y', errors='coerce')
        df_mes_atual = df[df['Data_dt'].dt.month == hoje.month]
        print(f"Filtrado para o mês: {hoje.month}") # Filtra o DataFrame para o mês atual

        destino = r"" # Caminho onde será salvo o arquivo Excel filtrado
        os.makedirs(destino, exist_ok=True) # Verifica se o diretório existe
        hoje_str = hoje.strftime('%d-%m-%Y') # Criação da variável com a data para dd/mm/aaaa
        nome_arquivo = os.path.join(destino, f"ocorrencias_{hoje_str}.xlsx") # Cria o nome do arquivo com a data do dia que foi criado
        print(f"Salvando em: {nome_arquivo}") 
        df_mes_atual.to_excel(nome_arquivo, index=False) # Salvar o DataFrame filtrado em um novo arquivo Excel com a data do dia que foi criado

        wb = load_workbook(nome_arquivo)
        ws = wb.active

        for col_nome, largura in colunas_para_ajustar.items(): # Loop para ajustar cada tamanho de coluna conforme a lista anteriormente criada
            col_idx = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == col_nome:
                    col_idx = idx
                    break
            if col_idx:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = largura # Organizar as colunas de acordo com o dicionário colunas_para_ajustar

        wb.save(nome_arquivo)
        print("Larguras ajustadas e arquivo salvo com sucesso.")

    except Exception as e:
        print(f"Erro: {e}") # Captura e exibe erros
    return nome_arquivo # Retorna o caminho do arquivo Excel gerado

def atualizar_powerbi_e_screenshot(): # Segunda função que atualiza o Power BI desktop e tira um print da área do dashboard
    caminho_pbi = r"" # Caminho do arquivo Power BI Desktop que será aberto
    coord_atualizar = (757, 102) # Coordenadas do botão de atualizar no Power BI Desktop
    area_vermelha = (77, 170, 1400, 760) # Área do print
    aba = (632,986) # Coordenadas da aba do dashboard no Power BI Desktop (caso haja mais de uma aba, é necessário clicar na aba correta)
    # Alterar as coordenadas conforme necessário para o seu monitor e resolução

    os.startfile(caminho_pbi) # Utiliza a função startfile do módulo os para abrir o Power BI Desktop
    print('Aguardando Power BI abrir...')
    time.sleep(15)

    pyautogui.moveTo(aba[0], aba[1], duration=0.5) # Move o mouse para a aba do dashboard
    pyautogui.click() # Clica na área onde o ponteiro está, no caso, a aba do dashboard
    time.sleep(1)
    pyautogui.moveTo(coord_atualizar[0], coord_atualizar[1], duration=0.5) # Move o mouse para o botão de atualizar
    pyautogui.click() # Clica na aba do dash para tirar print e clica no botão de atualizar
    print('Cliquei em Atualizar!')

    print('Aguardando atualização...')
    time.sleep(10)

    screenshot1 = pyautogui.screenshot()
    x, y, w, h = area_vermelha
    print1 = screenshot1.crop((x, y, x + w, y + h))

    destino = r"" # Caminho onde será salvo o print do dashboard
    hoje = datetime.now() # Criação da variável com a data atual
    hoje_str = hoje.strftime('%d-%m-%Y') # Criação da variável com o formato em dd/mm/aaaa
    caminho_semanal = os.path.join(destino, f"print_powerbi_semanal_{hoje_str}.png") # Cria o nome do arquivo com a data do dia que foi criado
    print1.save(caminho_semanal) # Salva o print do dashboard no caminho especificado
    print(f'Print salvo como {caminho_semanal}!')

    return caminho_semanal # Retorna o caminho do arquivo de print gerado

def enviar_email_relatorio(caminho_excel, caminho_print, destinatario, cc=None, bcc=None): # Terceira função que envia o e-mail com o relatório em anexo
    outlook = win32.Dispatch('outlook.application') # Cria uma instância do Outlook
    mail = outlook.CreateItem(0) # Cria um novo e-mail
    mail.To = destinatario # Define o destinatário do e-mail
    if cc: # Verifica se há destinatários em cópia, se tiver, os adiciona em cópia
        mail.CC = cc
    if bcc: # Verifica se há destinatários em cópia oculta, se tiver, os adiciona em cópia oculta
        mail.BCC = bcc
    mail.Subject = 'Relatório Score de segurança e meio ambiente' # Define o assunto do e-mail
    cid = 'print_dashboard' # Cria um ID para o conteúdo incorporado No corpo do e-mail

    # Define o corpo do e-mail em HTML, incluindo o print do dashboard como imagem incorporada
    mail.HTMLBody = f"""
    <p>Bom dia, segue o relatório de ocorrências:<br><br>
    <img src="cid:{cid}"><br>
    </p>
    """
    attachment = mail.Attachments.Add(os.path.abspath(caminho_print)) # Adiciona o print do dashboard como anexo
    attachment.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid) # Define o ID do conteúdo incorporado para a imagem
    mail.Attachments.Add(os.path.abspath(caminho_excel)) # Adiciona o arquivo Excel como anexo
    mail.Send() # Envia o email

if __name__ == "__main__": # Bloco principal que executa as funções na ordem correta
    caminho_excel = processar_planilha() # Executa a função para processar a planilha e retorna o caminho do arquivo Excel gerado
    caminho_print1 = atualizar_powerbi_e_screenshot() # Em seguida, executa a função para atualizar o Power BI e tirar um print do dashboard, retornando o caminho do print gerado
    destinatario = '' # Define o destinatário do e-mail
    cc = '' # Caso haja, define os destinatários em cópia
    bcc = None # Caso haja, define os destinatários em cópia oculta
    hoje = datetime.now() # Criação da variável com a data atual
    hoje_str = hoje.strftime('%d-%m-%Y') # Criação da variável com o formato em dd/mm/aaaa

    # Para definição dos caminhos, utilizei o nome padrão dos caminhos + a data do dia que foi criado, assim, não é necessário alterar os caminhos manualmente
    caminho_excel = r'' + hoje_str + '.xlsx' # Coloque o caminho do arquivo Excel gerado dentro das aspas
    caminho_print = r'' + hoje_str + '.png' # Coloque o caminho do print do dashboard gerado dentro das aspas
    enviar_email_relatorio(caminho_excel, caminho_print, destinatario, cc, bcc) # Executa a função para enviar o e-mail com o relatório em anexo

    pymsgbox.alert("Processo concluído com sucesso!", "Aviso") # Exibe uma mensagem de alerta informando que o processo foi concluído com sucesso